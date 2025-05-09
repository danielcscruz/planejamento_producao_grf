import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, timedelta
import os
from pathlib import Path

SETOR_ORDEM = [
    'PCP', 'Separação MP', 'Corte manual', 'Impressão',
    'Estampa', 'Corte laser', 'Costura', 'Arremate', 'Embalagem'
]

def obter_proximos_dias_uteis(data_inicio, dias_necessarios, calendario_path='data/_CALENDARIO.csv'):
    """
    Obtém uma lista de dias úteis a partir da data fornecida (inclusive).
    """
    # print(f"DEBUG - Buscando {dias_necessarios} dias úteis a partir de {data_inicio}")
    df_cal = pd.read_csv(calendario_path)
    
    # Converte a coluna DATA para datetime
    df_cal['DATA'] = pd.to_datetime(df_cal['DATA'], format="%d/%m/%Y")
    
    # Filtra apenas dias úteis
    dias_uteis = df_cal[df_cal['VALOR'] == 'UTIL']['DATA'].tolist()
    # print(f"DEBUG - Total de dias úteis no calendário: {len(dias_uteis)}")
    
    # Converte data_inicio para datetime se for string
    if isinstance(data_inicio, str):
        data_inicio = datetime.strptime(data_inicio, "%d/%m/%Y")
    
    # Usa a data exata fornecida
    data_proxima = data_inicio
    # print(f"DEBUG - Data para buscar próximo dia útil: {data_proxima}")
    
    # Encontra a posição do próximo dia útil a partir da data fornecida (inclusive)
    proxima_data_util_idx = 0
    encontrou = False
    for i, data in enumerate(dias_uteis):
        if data >= data_proxima:
            proxima_data_util_idx = i
            encontrou = True
            # print(f"DEBUG - Próximo dia útil encontrado: {data} no índice {i}")
            break
    
    if not encontrou:
        # print(f"DEBUG - ALERTA: Nenhum dia útil encontrado após {data_proxima}!")
        return []
    
    # Retorna os próximos dias úteis necessários
    dias_selecionados = dias_uteis[proxima_data_util_idx:proxima_data_util_idx + dias_necessarios]
    # print(f"DEBUG - Dias úteis selecionados: {dias_selecionados}")
    return dias_selecionados

def encontrar_coluna_por_data(ws, data):
    """
    Encontra a coluna na planilha que corresponde à data fornecida.
    Verifica tanto formato string quanto objeto datetime.
    """
    data_str = data.strftime("%d/%m/%Y")
    # print(f"DEBUG - Procurando coluna para a data: {data_str}")
    
    for col in range(8, ws.max_column + 1):  # Começando da coluna H (8)
        cell_value = ws.cell(row=2, column=col).value
        # Limita o debug para não sobrecarregar o console
        # if col % 10 == 0:  # Mostra debug a cada 10 colunas
            # print(f"DEBUG - Verificando células... coluna atual {col}")
        
        # Verifica se o valor da célula é uma data
        if isinstance(cell_value, datetime):
            # Compara o valor da data
            if cell_value.date() == data.date():
                # print(f"DEBUG - Data encontrada na coluna {col} (formato datetime)")
                return col
        elif isinstance(cell_value, str):
            # Tenta converter para datetime se for string
            try:
                # Tenta primeiro o formato DD/MM/YYYY
                cell_date = datetime.strptime(cell_value, "%d/%m/%Y")
                if cell_date.date() == data.date():
                    # print(f"DEBUG - Data encontrada na coluna {col} (formato string DD/MM/YYYY)")
                    return col
            except ValueError:
                try:
                    # Tenta o formato YYYY-MM-DD
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    if cell_date.date() == data.date():
                        # print(f"DEBUG - Data encontrada na coluna {col} (formato string YYYY-MM-DD)")
                        return col
                except ValueError:
                    # Ignora se não conseguir converter
                    pass
    
    # print(f"DEBUG - ALERTA: Data {data_str} não encontrada em nenhuma coluna da planilha!")
    return None

def salvar_nova_versao(caminho_original, workbook):
    """
    Salva a planilha em uma nova versão na pasta exp/
    """
    try:
        # Garantir que a pasta exp/ existe
        pasta_exp = Path("exp")
        pasta_exp.mkdir(exist_ok=True)
        
        # Obter o nome do arquivo original
        nome_arquivo = Path(caminho_original).name
        base_nome = Path(caminho_original).stem
        extensao = Path(caminho_original).suffix
        
        # Gerar um nome de arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        novo_nome = f"{base_nome}_{timestamp}{extensao}"
        novo_caminho = pasta_exp / novo_nome
        
        # Salvar a planilha no novo caminho
        workbook.save(str(novo_caminho))
        # print(f"DEBUG - Planilha salva como nova versão: {novo_caminho}")
        return str(novo_caminho)
    except Exception as e:
        # print(f"DEBUG - ERRO ao salvar nova versão da planilha: {e}")
        return None

def preencher_producao(ws: Worksheet, quantidade: int, setor: str, linha: int, 
                    corte, data_inicio=None, calendario_path='data/_CALENDARIO.csv', 
                    planilha_path=None, workbook=None, salvar: bool = True):
    """
    Preenche a produção a partir do setor especificado, propagando para os demais setores na ordem.
    
    Args:
        ws: Worksheet do openpyxl
        quantidade: Quantidade a ser produzida
        setor: Setor inicial
        linha: Linha na planilha
        data_inicio: Data de início (opcional)
        calendario_path: Caminho para o arquivo de calendário
        planilha_path: Caminho para o arquivo da planilha (para salvar nova versão)
        workbook: Objeto workbook do openpyxl (necessário para salvar)
    """
    # print(f"\nDEBUG - Iniciando planejamento: quantidade={quantidade}, setor={setor}, linha={linha}")    
    # Verifica se o arquivo de calendário existe
    # print(f"DEBUG - Verificando calendário: {calendario_path}")
    try:
        df_cal = pd.read_csv(calendario_path)
        # print(f"DEBUG - Calendário carregado com sucesso. {len(df_cal)} registros.")
    except Exception as e:
        # print(f"DEBUG - ERRO ao carregar calendário: {e}")
        return
    
    df_cal['DATA'] = pd.to_datetime(df_cal['DATA'], format="%d/%m/%Y")
    
    # Verifica se o setor é válido
    if setor not in SETOR_ORDEM:
        # print(f"DEBUG - ERRO: Setor '{setor}' não reconhecido.")
        raise ValueError(f"Setor '{setor}' não reconhecido. Deve ser um dos: {SETOR_ORDEM}")
    
    # Determina o índice do setor inicial na ordem de processamento
    setor_idx = SETOR_ORDEM.index(setor)
    setores_processar = SETOR_ORDEM[setor_idx:]
    # print(f"DEBUG - Setores a processar: {setores_processar}")
    
    # Se data_inicio não for fornecida, usa a data atual
    if data_inicio is None:
        data_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        # print(f"DEBUG - Data de início não fornecida, usando data atual: {data_inicio}")
    elif isinstance(data_inicio, str):
        data_inicio = datetime.strptime(data_inicio, "%d/%m/%Y")
        # print(f"DEBUG - Data de início convertida de string: {data_inicio}")
    
    # Para rastrear a data atual de processamento para cada setor
    data_atual = data_inicio
    ultimo_dia_usado = None
    # print(f"DEBUG - Data inicial para processamento: {data_atual}")
    primeiro_dia_usado = None
    delay = 0
    # Processa cada setor na ordem
    for i, setor_nome in enumerate(setores_processar):

        # print(f"\nDEBUG - Processando setor: {setor_nome} (índice {i})")
        
        # print(f"\nDEBUG - Comparando corte e setor: {setor_nome} e {corte}")
        if setor_nome == 'Corte manual' and corte == 'Corte laser':
            # print(f"\nDEBUG - Break: {setor_nome} e {corte}")
            continue
        if setor_nome == 'Corte laser' and corte == 'Corte manual':
            # print(f"\nDEBUG - Break: {setor_nome} e {corte}")
            continue
        
        # Se não for o primeiro setor, avança para o próximo dia útil
        if i > 0:
            # print(f"DEBUG - Processando próximo setor, avançando para o próximo dia útil")
            
            # CORREÇÃO: Garantir que avançamos pelo menos um dia para o próximo setor
            # Avança um dia para garantir que estamos no dia seguinte
            if ultimo_dia_usado:
                # Use o último dia que foi realmente usado pelo setor anterior
                data_proxima = ultimo_dia_usado + timedelta(days=1)
            else:
                # Fallback se não tivermos registro do último dia usado
                data_proxima = data_atual + timedelta(days=1)
                
            # print(f"DEBUG - Avançando para o dia seguinte: {data_proxima}")
            
            # Agora buscamos o próximo dia útil a partir desse dia seguinte
            proximos_dias = obter_proximos_dias_uteis(data_proxima, 1, calendario_path)
            if proximos_dias:
                data_atual = proximos_dias[0]
                # print(f"DEBUG - Próximo dia útil encontrado: {data_atual}")
            else:
                # print(f"DEBUG - ALERTA: Nenhum dia útil encontrado após {data_proxima}!")
                # Avança um dia como fallback
                data_atual = data_proxima
        
        # Resetamos o último dia usado para este novo setor
        ultimo_dia_usado = None
        
        # Calcula a linha do setor atual (linha do pedido + offset do setor)
        linha_setor = linha + SETOR_ORDEM.index(setor_nome) + 1
        # print(f"DEBUG - Linha do setor {setor_nome}: {linha_setor}")
        
        # Obtém o limite de produção diário para este setor
        linha_limite = SETOR_ORDEM.index(setor_nome) + 3  # +3 para corresponder à posição na planilha
        try:
            valor_limite_cell = ws.cell(row=linha_limite, column=5).value
            # print(f"DEBUG - Célula de limite na linha {linha_limite}, coluna 5: {valor_limite_cell}")
            valor_limite_max = int(valor_limite_cell) if valor_limite_cell is not None else 0
        except (ValueError, TypeError) as e:
            # print(f"DEBUG - Erro ao converter limite: {e}")
            valor_limite_max = 0
        
        # print(f"DEBUG - >> Limite máximo diário para o setor {setor_nome}: {valor_limite_max }")
        
        qtd_restante = quantidade
        
        # CORREÇÃO: Enquanto houver quantidade restante, continue processando o setor atual
        while qtd_restante > 0:
            # Calcula quantos dias serão necessários para processar toda a quantidade restante
            if valor_limite_max > 0:
                dias_necessarios = (qtd_restante + valor_limite_max - 1) // valor_limite_max  # Arredonda para cima
            else:
                dias_necessarios = 1  # Se não há limite, assume que tudo pode ser feito em um dia
            
            # Limita o número de dias buscados por vez para evitar carregar toda a tabela
            dias_por_iteracao = min(dias_necessarios, 30)  # Busca no máximo 30 dias por vez
            
            # print(f"DEBUG - Buscando {dias_por_iteracao} dias úteis para processar {qtd_restante} unidades")
            
            # Obtém os próximos dias úteis necessários
            dias_uteis = obter_proximos_dias_uteis(data_atual, dias_por_iteracao, calendario_path)
            
            if not dias_uteis:
                # print(f"DEBUG - ALERTA: Nenhum dia útil encontrado para o setor {setor_nome}!")
                # CORREÇÃO: Se não encontrar dias úteis, saímos do loop deste setor e continuamos
                # para evitar loop infinito, mas com um alerta claro
                # print(f"DEBUG - ERRO CRÍTICO: Não foi possível alocar {qtd_restante} unidades para o setor {setor_nome}!")
                break
            
            # print(f"DEBUG - Distribuindo produção para o setor {setor_nome}")
            # Distribui a produção pelos dias úteis
            for dia_util in dias_uteis:
                # Encontra a coluna correspondente à data
                col = encontrar_coluna_por_data(ws, dia_util)
                
                if col is None:
                    # print(f"DEBUG - Pulando dia {dia_util} pois não foi encontrado na planilha")
                    continue
                
                # Debugging: Print the values being read from the cells
                valor_planejado = 0  # Initialize the variable
                for row in range(12, linha_setor):
                    cell_value = ws.cell(row=row, column=col).value
                    setor_cell_value = ws.cell(row=row, column=7).value  # Coluna G é a 7ª coluna
                    if setor_cell_value == setor_nome:
                        valor_planejado += int(cell_value or 0)  # Calculate the sum of the values
            
                # print(f"DEBUG - >>>>Soma calculada para valor_planejado: {valor_planejado}")
                # print(f"DEBUG - >> Valor já planejado até a linha {linha_setor}, coluna {col}: {valor_planejado}")

                # Calcula o limite disponível para o dia atual
                valor_limite = max(0, valor_limite_max - valor_planejado)
                # print(f"DEBUG - Limite diário disponível para o setor {setor_nome}: {valor_limite}")

                # Determina quanto produzir neste dia
                # producao_dia = min(valor_limite, qtd_restante) if valor_limite > 0 else qtd_restante
                
                if qtd_restante > valor_limite:
                    producao_dia = valor_limite
                else:
                    producao_dia = qtd_restante

                # Registra a produção na planilha
                # print(f"DEBUG - Preenchendo: linha={linha_setor}, coluna={col}, valor={producao_dia}")
                # current_value = ws.cell(row=linha_setor, column=col).value
                # print(f"DEBUG - Valor atual na célula: {current_value}")
                
                try:
                    ws.cell(row=linha_setor, column=col, value=producao_dia)
                    # print(f"DEBUG - Produção registrada com sucesso: {producao_dia} unidades")
                    
                    # Registra este dia como o último usado pelo setor atual
                    ultimo_dia_usado = dia_util
                    if primeiro_dia_usado is None and producao_dia > 0:
                        primeiro_dia_usado = dia_util
                        print(f"DEBUG - primeiro dia: {primeiro_dia_usado}")

                    if producao_dia == 0:
                        delay += 1

                except Exception as e:
                    print(f"DEBUG - ERRO ao registrar produção: {e}")
                
                qtd_restante -= producao_dia
                # print(f"DEBUG - Quantidade restante: {qtd_restante}")
                
                # Se toda a quantidade foi distribuída, para
                if qtd_restante <= 0:
                    # print(f"DEBUG - Toda a quantidade foi distribuída para o setor {setor_nome}")
                    break
            
            # CORREÇÃO: Se ainda houver quantidade restante, avança a data atual para continuar
            # buscando mais dias úteis a partir do último dia utilizado + 1
            if qtd_restante > 0:
                if ultimo_dia_usado:
                    data_atual = ultimo_dia_usado + timedelta(days=1)
                else:
                    data_atual = dias_uteis[-1] + timedelta(days=1)
                # print(f"DEBUG - Ainda restam {qtd_restante} unidades. Avançando para {data_atual} para buscar mais dias úteis.")
            
            # CORREÇÃO: Se foram processados todos os dias disponíveis e ainda há qtd_restante,
            # mas não conseguimos alocar nada (porque não existem colunas na planilha para essas datas),
            # devemos sair do loop para evitar um loop infinito
            if all(encontrar_coluna_por_data(ws, dia) is None for dia in dias_uteis) and len(dias_uteis) > 0:
                # print(f"DEBUG - ALERTA: Nenhuma das datas obtidas está disponível na planilha. Saindo do loop.")
                # print(f"DEBUG - ERRO CRÍTICO: Não foi possível alocar {qtd_restante} unidades para o setor {setor_nome}!")
                break
    
    # print("DEBUG - Planejamento de produção concluído")
    # print(f"{corte}")
    
    # Salvar a planilha em uma nova versão se um caminho e o workbook foram fornecidos
    if planilha_path and workbook:
        if salvar:
            salvar_nova_versao(planilha_path, workbook)

    return primeiro_dia_usado, ultimo_dia_usado, delay