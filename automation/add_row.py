from openpyxl import load_workbook
from datetime import datetime
from InquirerPy import inquirer
from automation.fill_production import preencher_producao


def validar_data_input(data_str):
    try:
        return datetime.strptime(data_str, '%d/%m/%Y')
    except ValueError:
        return None


def adicionar_nova_linha(arquivo_path):
    wb = load_workbook(arquivo_path)
    ws = wb.active

    # Encontra a primeira linha vazia após o cabeçalho (a partir da linha 12)
    linha = 12
    while ws.cell(row=linha, column=1).value:
        linha += 1

    print("\n📥 Preencha as informações para adicionar uma nova linha à planilha:")

    pedido = input("🆔 Pedido: ").strip()

    while True:
        entrega_str = input("📅 Data de Entrega (DD/MM/AAAA): ").strip()
        entrega = validar_data_input(entrega_str)
        if entrega:
            break
        print("❌ Data inválida. Tente novamente no formato DD/MM/AAAA.")

    cliente = input("👤 Cliente: ").strip()
    produto = input("📦 Produto: ").strip()

    while True:
        try:
            quantidade = int(input("🔢 Quantidade: ").strip())
            break
        except ValueError:
            print("❌ Quantidade inválida. Digite um número inteiro.")

    # 🔽 Seleção do tipo de corte usando InquirerPy
    tipo_corte = inquirer.select(
        message="🛠️ Selecione o tipo de corte:",
        choices=["Corte Manual", "Corte a Laser"],
        default="Corte Manual",
    ).execute()

    # Preenche os dados na planilha
    ws.cell(row=linha, column=1, value=pedido)
    ws.cell(row=linha, column=2, value=entrega)
    ws.cell(row=linha, column=3, value=cliente)
    ws.cell(row=linha, column=4, value=produto)
    ws.cell(row=linha, column=5, value=quantidade)

    preencher_producao(ws, quantidade=quantidade, setor='PCP', linha=linha)


    wb.save(arquivo_path)
    wb.close()

    print(f"\n✅ Nova linha adicionada com sucesso na linha {linha} da planilha: {arquivo_path}")
    return tipo_corte
