import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
from automation.core.constants import SETOR_ORDEM

def gerar_relatorio_arquivo(arquivo_path: str):
    print("Iniciou função de exportar arquivo")
    print(f"Arquivo base: {arquivo_path}")

    arquivo_full_path = os.path.join('exp', os.path.basename(arquivo_path))
    print(f"Arquivo completo com rota: {arquivo_full_path}")

    setores = SETOR_ORDEM

    nome_base = os.path.splitext(os.path.basename(arquivo_full_path))[0]
    output_dir = os.path.join(os.path.dirname(arquivo_full_path), nome_base)
    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(arquivo_full_path, data_only=True)
    ws = wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    df_base = pd.DataFrame(data)

    for setor in setores:
        print(f"Processando setor: {setor}")
        df_setor = df_base.copy()
        df_setor.drop(index=range(2, 12), inplace=True, errors='ignore')
        df_setor.iloc[0, 6:] = df_setor.iloc[0, 6:].fillna("")

        df_head = df_setor.iloc[1:2]
        df_body = df_setor.iloc[2:]

        df_filtered = df_body[
            (df_body[6] == setor) &
            (df_body[0].notna() | df_body[4].notna())
        ]

        df_final = pd.concat([df_head, df_filtered])

        # Resetar índices para facilitar o processamento
        df_final = df_final.reset_index(drop=True)
        
        # Criar um novo DataFrame convertendo todas as colunas para evitar warnings
        dados_convertidos = []
        
        for row_idx in range(len(df_final)):
            linha_convertida = []
            for col_idx in range(df_final.shape[1]):
                valor = df_final.iat[row_idx, col_idx]
                
                # Se for uma coluna de data (a partir do índice 7) e for datetime, converter
                if col_idx >= 7 and isinstance(valor, datetime) and pd.notna(valor):
                    linha_convertida.append(valor.strftime('%d/%m'))
                else:
                    linha_convertida.append(valor)
            
            dados_convertidos.append(linha_convertida)
        
        # Criar novo DataFrame com os dados convertidos
        df_final = pd.DataFrame(dados_convertidos)

        output_filename = f"planejamento_{setor}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        df_final.to_excel(output_path, index=False, header=False)

        print(f"  ✔ Arquivo salvo: {output_path}")

    print("\n" + "="*50)
    print("✅ PROCESSO CONCLUÍDO COM SUCESSO!")
    print(f"📁 Todos os relatórios foram salvos em: {output_dir}")
    print("="*50)
    input("\nPressione ENTER para voltar ao menu inicial...")