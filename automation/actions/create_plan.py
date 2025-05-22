from openpyxl import load_workbook
import pandas as pd
import os
from tabulate import tabulate
from InquirerPy import inquirer

from automation.core.production_planner import preencher_producao  
from automation.core.excel_utils import atualizar_limites_maximos, atualizar_celulas_limite, obter_carga_producao
from automation.core.constants import DEFAULT_CONFIG_PATH


def criar_novo_plano(df_priorizado: pd.DataFrame):
    total = len(df_priorizado)
    arquivo_path = "model/planejamento.xlsx"

    wb = load_workbook(arquivo_path)
    ws = wb.active

    ultimo_dia_list = []
    primeiro_dia_list = []
    delay_list = []

    # Gera a lista de limites máximos
    max_list = atualizar_limites_maximos(config_path=DEFAULT_CONFIG_PATH)

    carga_prod = obter_carga_producao(config_path=DEFAULT_CONFIG_PATH)

    # Atualizando os valores da lista com base na porcentagem da carga
    max_list_carga = [valor * (carga_prod / 100) for valor in max_list]
    
    # Atualiza as células [E3:E12] na planilha
    atualizar_celulas_limite(ws, max_list_carga)



    for index, row in df_priorizado.iterrows():
        linha = 13
        while ws.cell(row=linha, column=1).value:
            linha += 1

        # Pegar valores da tabela
        pedido = row.get("PEDIDO")
        entrega = row.get("ENTREGA")
        cliente = row.get("CLIENTE")
        produto = row.get("PRODUTO")
        quantidade= row.get("QUANTIDADE")
        corte = row.get("CORTE")
        setor=row.get("SETOR")
        inicio_data=row.get("INICIO")
        
        # Preenche os dados na planilha
        ws.cell(row=linha, column=1, value=pedido)
        ws.cell(row=linha, column=2, value=entrega)
        ws.cell(row=linha, column=3, value=cliente)
        ws.cell(row=linha, column=4, value=produto)
        ws.cell(row=linha, column=5, value=quantidade)


        if not corte:
            print(f"[Linha {index}] TIPO DE CORTE não especificado. Pulando.")
            ultimo_dia_list.append(None)
            primeiro_dia_list.append(None)
            delay_list.append(None)
            continue

        try:
            salvar = (index == total - 1)  # só salva no último
            print(f"[Linha {index}] Iniciando preenchimento para tipo de corte: {corte}")

            primeiro_dia, ultimo_dia, delay = preencher_producao(
                ws=ws, 
                df_priorizado=df_priorizado, 
                quantidade=quantidade, 
                setor=setor, 
                linha=linha, 
                calendario_path="data/_CALENDARIO.csv", 
                planilha_path=arquivo_path, 
                workbook=wb,
                data_inicio=inicio_data,
                corte=corte, 
                salvar=salvar
            )

            # Padroniza a data para o formato DD/MM/AAAA
            if ultimo_dia:
                ultimo_dia = ultimo_dia.strftime('%d/%m/%Y')
            if primeiro_dia:
                primeiro_dia = primeiro_dia.strftime('%d/%m/%Y')

            primeiro_dia_list.append(primeiro_dia)
            ultimo_dia_list.append(ultimo_dia)
            delay_list.append(delay)
            print(f"[Linha {index}] Preenchimento concluído.\n")
        except Exception as e:
            print(f"[Linha {index}] Erro ao preencher produção: {e}")
            ultimo_dia_list.append(None)
            primeiro_dia_list.append(None)
            delay_list.append(None)

    df_priorizado["PRIMEIRO DIA"] = primeiro_dia_list   
    df_priorizado["ULTIMO DIA"] = ultimo_dia_list
    df_priorizado["DELAY"] = delay_list


    return df_priorizado, carga_prod