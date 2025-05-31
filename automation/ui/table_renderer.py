import pandas as pd
import os
from datetime import datetime
from tabulate import tabulate
from openpyxl import load_workbook

def validar_data_input(data_str):
    try:
        return datetime.strptime(data_str, '%d/%m/%Y')
    except ValueError:
        return None

def processar_tabela(file_choice):
    print("Lendo o arquivo inicial...")
    arquivo_destino = f"{file_choice}.xlsx" if not str(file_choice).endswith('.xlsx') else file_choice
    print(f"\nAbrindo planilha: {arquivo_destino}")
    
    wb = load_workbook(arquivo_destino)
    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
        if row and "Pedido" in row:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Coluna 'Pedido' não encontrada!")
    
    print(f"-> Cabeçalho encontrado na linha {header_row + 1}")
    
    df = pd.read_excel(arquivo_destino, header=header_row)
    df = df[df["Pedido"].notna()]
    df_unique = df.drop_duplicates(subset="Pedido", keep="first")
    
    colunas_desejadas = ["Pedido", "Entrega", "Cliente", "Produto", "QTD", "Tipo de Corte", "Data Inicio", "Setor"]
    df_formatado = df_unique[colunas_desejadas].copy()
    df_formatado.columns = ['PEDIDO', 'ENTREGA', 'CLIENTE', 'PRODUTO', 'QUANTIDADE', 'CORTE', 'INICIO', 'SETOR']
    
    # Corrigindo a formatação do PEDIDO - remove apenas decimais desnecessários
    def limpar_pedido(valor):
        valor_str = str(valor).strip()
        # Remove apenas se terminar com .0 ou .00 (decimais desnecessários)
        # Usa regex mais específica que só remove ponto seguido de zeros no final
        import re
        return re.sub(r'\.0+$', '', valor_str)
    
    df_formatado['PEDIDO'] = df_formatado['PEDIDO'].apply(limpar_pedido)

    # Validação de datas (ENTREGA e DATA)
    for col in ['ENTREGA', 'INICIO']:
        for idx, valor in df_formatado[col].items():
            if isinstance(valor, (datetime, pd.Timestamp)):
                continue
            if str(valor).strip() in ['', 'nan', 'NaT']:
                print(f"\n⚠️ Valor inválido detectado na coluna '{col}' na linha {idx + header_row + 2} (valor: '{valor}')")
                while True:
                    nova_data_str = input(f"Digite uma nova data para '{valor}' no formato DD/MM/AAAA: ")
                    nova_data = validar_data_input(nova_data_str)
                    if nova_data:
                        ws.cell(row=idx + header_row + 2, column=colunas_desejadas.index(col.capitalize()) + 1).value = nova_data
                        print(f"✔️ Corrigido para {nova_data_str}")
                        break
                    else:
                        print("❌ Formato inválido. Tente novamente.")
    
    wb.save(arquivo_destino)
    wb.close()
    
    # Recarrega para aplicar as correções
    df = pd.read_excel(arquivo_destino, header=header_row)
    df = df[df["Pedido"].notna()]
    df_unique = df.drop_duplicates(subset="Pedido", keep="first")
    df_formatado = df_unique[colunas_desejadas].copy()
    df_formatado.columns = ['PEDIDO', 'ENTREGA', 'CLIENTE', 'PRODUTO', 'QUANTIDADE', 'CORTE', 'INICIO', 'SETOR']
    
    # Aplica a mesma correção após recarregar
    df_formatado['PEDIDO'] = df_formatado['PEDIDO'].apply(limpar_pedido)

    # Converte colunas de data
    df_formatado['ENTREGA'] = pd.to_datetime(df_formatado['ENTREGA'], errors='coerce')
    df_formatado['INICIO'] = pd.to_datetime(df_formatado['INICIO'], errors='coerce')

    # Validação da coluna 'CORTE'
    df_formatado['CORTE'] = df_formatado['CORTE'].astype(str).str.strip().str.lower().map(
        lambda x: 'Laser' if x == 'laser' else 'Manual'
    )

    # Validação da coluna 'SETOR'
    opcoes_validas_inicio = [
        'PCP', 'Separação MP', 'Corte manual', 'Impressão',
        'Estampa', 'Corte laser', 'Costura', 'Arremate', 'Embalagem'
    ]
    opcoes_validas_normalizadas = [op.lower() for op in opcoes_validas_inicio]
    df_formatado['SETOR'] = df_formatado['SETOR'].astype(str).str.strip().map(
        lambda x: next((op for op in opcoes_validas_inicio if op.lower() == x.lower()), 'PCP')
    )

    # Exibição final
    os.system('cls' if os.name == 'nt' else 'clear')
    df_formatado['ENTREGA'] = df_formatado['ENTREGA'].dt.strftime('%d/%m/%Y')
    df_formatado['INICIO'] = df_formatado['INICIO'].dt.strftime('%d/%m/%Y')
    
    print("\n✅ Dados formatados com sucesso:\n")
    print(tabulate(df_formatado, headers='keys', tablefmt='grid', showindex=False))

    produtos_unicos = df_formatado['PRODUTO'].unique().tolist()
    return df_formatado, produtos_unicos