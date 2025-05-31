"""
Módulo principal para planejamento de produção com fluxo contínuo.
"""

from datetime import datetime, timedelta
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from InquirerPy import inquirer
import csv
import unicodedata

from automation.core.constants import SETOR_ORDEM, DEFAULT_CALENDARIO_PATH, DEFAULT_CONFIG_PATH, obter_valor_parametro
from automation.core.calendar_utils import obter_proximos_dias_uteis
from automation.core.excel_utils import encontrar_coluna_por_data, calcular_producao_planejada, obter_limite_producao
from automation.core.file_utils import salvar_nova_versao


def obter_delta_dias_estampa(config_path=DEFAULT_CONFIG_PATH):
    """
    Obtém o valor de DELTA_DIAS_ESTAMPA do arquivo de configuração CSV.

    Args:
        config_path (str): Caminho para o arquivo de configuração CSV.

    Returns:
        int: Valor de DELTA_DIAS_ESTAMPA, ou 5 como padrão se não encontrado.
    """
    try:
        with open(config_path, 'r') as config_file:
            reader = csv.DictReader(config_file)
            for row in reader:
                if row['PARAMETRO'] == 'DELTA_DIAS_ESTAMPA':
                    return int(row['VALOR'])  # Converte o valor para inteiro
    except FileNotFoundError:
        print(f"Arquivo de configuração não encontrado: {config_path}. Usando valor padrão (5).")
    except ValueError:
        print(f"Erro ao converter o valor de DELTA_DIAS_ESTAMPA para inteiro. Usando valor padrão (5).")
    except Exception as e:
        print(f"Erro inesperado ao carregar o arquivo de configuração: {e}. Usando valor padrão (5).")
    return 5

def preencher_producao(
        ws: Worksheet,
        df_priorizado: pd.DataFrame,
        quantidade: int, 
        setor: str, 
        linha: int, 
        corte, 
        data_inicio=None, 
        calendario_path=DEFAULT_CALENDARIO_PATH, 
        planilha_path=None, 
        workbook=None, 
        salvar: bool = True,
        priorizar_estampa: bool = None
    ):
    """
    Preenche a produção a partir do setor especificado, com fluxo contínuo entre setores.
    
    Args:
        ws: Worksheet do openpyxl
        df_priorizado: DataFrame do pandas com os dados de produção
        quantidade: Quantidade a ser produzida
        setor: Setor inicial
        linha: Linha na planilha
        corte: Tipo de corte ('Corte manual' ou 'Corte laser')
        data_inicio: Data de início (opcional)
        calendario_path: Caminho para o arquivo de calendário
        planilha_path: Caminho para o arquivo da planilha (para salvar nova versão)
        workbook: Objeto workbook do openpyxl (necessário para salvar)
        salvar: Se True, salva uma nova versão da planilha
        priorizar_estampa: Se True, prioriza terças e quintas para o setor Estampa
        
    Returns:
        tuple: (primeiro_dia_usado, ultimo_dia_usado, delay)
    """
    # Verifica se o setor é válido
    if setor not in SETOR_ORDEM:
        raise ValueError(f"Setor '{setor}' não reconhecido. Deve ser um dos: {SETOR_ORDEM}")
    
    # Determina o índice do setor inicial na ordem de processamento
    setor_idx = SETOR_ORDEM.index(setor)
    setores_processar = SETOR_ORDEM[setor_idx:]
    print(f"DEBUG - setor_idx: {setor_idx} | setores_processar: {setores_processar}")
    
    # Se data_inicio não for fornecida, usa a data atual
    if data_inicio is None:
        data_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    elif isinstance(data_inicio, str):
        data_inicio = datetime.strptime(data_inicio, "%d/%m/%Y")
    
    priorizar_estampa_value = obter_valor_parametro('PRIORIDADE_ESTAMPA')
    if priorizar_estampa_value == 'Sim':
        priorizar_estampa = True
    else:
        priorizar_estampa = False
    
    # Obtendo informações do pedido atual para referência
    pedido = ws.cell(row=linha, column=1).value
    entrega = ws.cell(row=linha, column=2).value
    cliente = ws.cell(row=linha, column=3).value
    produto = ws.cell(row=linha, column=4).value

    print(f"\n--- Informações do Pedido ---")
    print(f"Linha: {linha}")
    print(f"Pedido: {pedido}")
    print(f"Entrega: {entrega}")
    print(f"Cliente: {cliente}")
    print(f"Produto: {produto}")
    print(f"Quantidade: {quantidade}")
    print(f"Tipo de Corte: {corte}")
    print(f"------------------------\n")
    
    # Executa o planejamento com fluxo contínuo
    primeiro_dia_usado, ultimo_dia_usado, delay = _processar_fluxo_continuo(
        ws, setores_processar, quantidade, linha, data_inicio, 
        calendario_path, corte, priorizar_estampa
    )
    
    # Salvar a planilha em uma nova versão se um caminho e o workbook foram fornecidos
    if planilha_path and workbook and salvar:
        salvar_nova_versao(planilha_path, workbook)

    return primeiro_dia_usado, ultimo_dia_usado, delay

def _processar_fluxo_continuo(
        ws: Worksheet, setores_processar: list, quantidade_total: int, 
        linha: int, data_inicio: datetime, calendario_path: str, 
        corte: str, priorizar_estampa: bool
    ):
    """
    Processa todos os setores com fluxo contínuo, onde a produção de cada dia 
    fica disponível para o próximo setor no dia seguinte.
    """
    # Inicialização
    primeiro_dia_usado = None
    ultimo_dia_usado = None
    delay = 0
    
    # Filtrar setores válidos baseado no tipo de corte
    setores_validos = []
    for setor in setores_processar:
        if not _deve_pular_setor(setor, corte):
            setores_validos.append(setor)
    
    print(f"🔧 Setores a processar (após filtro de corte): {setores_validos}")
    
    # Controle de estoque entre setores: {setor: quantidade_disponivel}
    quantidade_disponivel = {}
    
    # Inicializar controles
    for i, setor in enumerate(setores_validos):
        if i == 0:
            # Primeiro setor tem toda a quantidade para produzir
            quantidade_disponivel[setor] = quantidade_total
        else:
            # Outros setores começam sem nada para produzir
            quantidade_disponivel[setor] = 0
    
    # Obter configurações de cada setor
    config_setores = {}
    for setor in setores_validos:
        config_setores[setor] = _obter_config_setor(setor, ws)
    
    # Obter dias úteis para um período amplo (90 dias)
    dias_uteis = obter_proximos_dias_uteis(data_inicio, 90, calendario_path)
    
    if not dias_uteis:
        print("⚠ Nenhum dia útil encontrado!")
        return primeiro_dia_usado, ultimo_dia_usado, delay
    
    # Controle de produção acumulada por setor
    producao_acumulada = {setor: 0 for setor in setores_validos}
    
    print(f"📊 Iniciando processamento - Quantidade total: {quantidade_total}")
    print(f"🎯 Setores válidos: {setores_validos}")
    
    # Processar dia a dia
    for idx_dia, dia_atual in enumerate(dias_uteis):
        print(f"\n📅 Processando dia: {dia_atual.strftime('%d/%m/%Y')}")
        dia_teve_producao = False
        producao_dia = {}
        
        # Transferir estoque ANTES do processamento do dia (exceto no primeiro dia)
        if idx_dia > 0:
            _transferir_estoque_dia_anterior(setores_validos, producao_dia_anterior, quantidade_disponivel)
        
        # Processar cada setor neste dia
        for i, setor in enumerate(setores_validos):
            producao_dia[setor] = 0
            
            # Verificar se há material disponível para produzir
            if quantidade_disponivel[setor] <= 0:
                print(f"  [{setor}] Sem material disponível ({quantidade_disponivel[setor]})")
                continue
            
            # Verificar priorização de estampa (só terças e quintas)
            if setor == 'Estampa' and priorizar_estampa:
                if dia_atual.weekday() not in [1, 3]:  # 1=terça, 3=quinta
                    print(f"  [{setor}] Dia não prioritário para estampa (apenas terças e quintas)")
                    delay += 1
                    continue
            
            # Processar produção do setor neste dia
            producao_realizada = _processar_setor_dia(
                ws, setor, dia_atual, linha, quantidade_disponivel[setor], 
                config_setores[setor]
            )
            
            if producao_realizada > 0:
                dia_teve_producao = True
                producao_dia[setor] = producao_realizada
                
                # Registrar primeiro e último dia usado
                if primeiro_dia_usado is None:
                    primeiro_dia_usado = dia_atual
                ultimo_dia_usado = dia_atual
                
                # Atualizar quantidade disponível do setor atual
                quantidade_disponivel[setor] -= producao_realizada
                producao_acumulada[setor] += producao_realizada
                
                print(f"  ✔[{setor}] {producao_realizada} unidades produzidas (disponível: {quantidade_disponivel[setor]}, acumulado: {producao_acumulada[setor]})")
            else:
                print(f"  [-] [{setor}] Nenhuma produção (material: {quantidade_disponivel[setor]})")
        
        # Salvar produção do dia para transferência no próximo dia
        producao_dia_anterior = producao_dia.copy()
        
        # Verificar se toda a produção foi concluída
        if producao_acumulada[setores_validos[-1]] >= quantidade_total:
            print(f"✅ Toda a produção foi concluída até {dia_atual.strftime('%d/%m/%Y')}")
            print(f"📊 Produção final do último setor: {producao_acumulada[setores_validos[-1]]}/{quantidade_total}")
            break
        
        # Incrementar delay se nenhum setor produziu neste dia
        if not dia_teve_producao:
            delay += 1
        
        # Debug: mostrar status atual
        print(f"  📊 Status: {[f'{setor}:{quantidade_disponivel[setor]}' for setor in setores_validos]}")
    
    return primeiro_dia_usado, ultimo_dia_usado, delay

def _obter_config_setor(setor_nome: str, ws: Worksheet):
    """Obtém as configurações de um setor específico."""
    # Lambda function para transformar a string
    formatar_parametro = lambda setor_nome: "MAX_" + unicodedata.normalize('NFKD', setor_nome).encode('ASCII', 'ignore').decode('ASCII').upper().replace(" ", "_")
    
    setor_formatado = formatar_parametro(setor_nome)
    valor_parametro_max = float(obter_valor_parametro(setor_formatado))
    valor_setup = float(obter_valor_parametro('SETUP'))
    setup = valor_parametro_max * valor_setup / 100
    
    # Obter limite de produção da planilha
    linha_limite = SETOR_ORDEM.index(setor_nome) + 3
    valor_limite_max = obter_limite_producao(ws, linha_limite)
    
    return {
        'limite_max': valor_limite_max,
        'setup': setup,
        'sem_limite': setor_nome in ["PCP", "Separação MP"]
    }

def _deve_pular_setor(setor_nome: str, corte: str):
    """Verifica se um setor deve ser pulado baseado no tipo de corte."""
    return ((setor_nome == 'Corte manual' and corte.lower() == 'laser') or 
            (setor_nome == 'Corte laser' and corte.lower() == 'manual'))

def _processar_setor_dia(ws: Worksheet, setor: str, dia: datetime, linha: int, 
                        quantidade_disponivel: int, config_setor: dict):
    """
    Processa a produção de um setor em um dia específico.
    
    Returns:
        int: Quantidade efetivamente produzida
    """
    if quantidade_disponivel <= 0:
        return 0
        
    # Calcular linha do setor na planilha
    linha_setor = linha + SETOR_ORDEM.index(setor) + 1
    
    # Encontrar coluna da data
    col = encontrar_coluna_por_data(ws, dia)
    if col is None:
        print(f"    ⚠ [{setor}] Coluna não encontrada para {dia.strftime('%d/%m/%Y')}")
        return 0
    
    # Calcular produção já planejada
    valor_planejado = calcular_producao_planejada(ws, setor, col, linha_setor)
    
    # Calcular limite disponível
    if config_setor['sem_limite']:
        # Setores sem limite diário - pode produzir tudo que tem disponível
        limite_disponivel = quantidade_disponivel
        print(f"    🔄 [{setor}] Setor sem limite - pode produzir: {quantidade_disponivel}")
    else:
        limite_disponivel = max(0, config_setor['limite_max'] - valor_planejado)
        print(f"    📊 [{setor}] Limite máx: {config_setor['limite_max']}, já planejado: {valor_planejado}, disponível: {limite_disponivel}")
    
    # Verificar setup mínimo para setores com limite
    if not config_setor['sem_limite'] and limite_disponivel > 0 and limite_disponivel < config_setor['setup']:
        print(f"    ⚠ [{setor}] Limite disponível ({limite_disponivel}) menor que setup ({config_setor['setup']})")
        return 0
    
    # Determinar quanto produzir
    if config_setor['sem_limite']:
        producao_dia = quantidade_disponivel
    else:
        producao_dia = min(limite_disponivel, quantidade_disponivel)
    
    # Registrar na planilha
    if producao_dia > 0:
        try:
            valor_atual = ws.cell(row=linha_setor, column=col).value or 0
            novo_valor = valor_atual + producao_dia
            ws.cell(row=linha_setor, column=col, value=novo_valor)
            print(f"    ✅ [{setor}] Registrado na planilha: {producao_dia} (total na célula: {novo_valor})")
        except Exception as e:
            print(f"    ⚠ [{setor}] Erro ao registrar produção: {e}")
            return 0
    else:
        print(f"    [-] [{setor}] Nenhuma produção registrada")
    
    return producao_dia

def _transferir_estoque_dia_anterior(setores_validos: list, producao_dia_anterior: dict, quantidade_disponivel: dict):
    """
    Transfere a produção do dia anterior para o próximo setor.
    """
    for i in range(len(setores_validos) - 1):
        setor_atual = setores_validos[i]
        proximo_setor = setores_validos[i + 1]
        
        # Verificar se há produção do setor atual para transferir
        if setor_atual in producao_dia_anterior and producao_dia_anterior[setor_atual] > 0:
            quantidade_transferir = producao_dia_anterior[setor_atual]
            
            # Transferir para o próximo setor
            quantidade_disponivel[proximo_setor] += quantidade_transferir
            
            print(f"    📦 Transferido: {quantidade_transferir} unidades de '{setor_atual}' para '{proximo_setor}'")

def _transferir_estoque_entre_setores(setores_processar: list, estoque_intermediario: dict, 
                                    quantidade_pendente: dict, dia_producao: datetime, corte: str):
    """
    FUNÇÃO REMOVIDA - Substituída por _transferir_estoque_dia_anterior
    """
    pass

def _processar_setor(
        ws: Worksheet, setor_nome: str, quantidade: int, linha: int, 
        data_atual: datetime, calendario_path: str, 
        primeiro_dia_usado=None, ultimo_dia_usado=None, delay=0,
        priorizar_estampa_terca_quinta=False
    ):
    """
    FUNÇÃO MANTIDA PARA COMPATIBILIDADE - MAS NÃO É MAIS USADA NO FLUXO CONTÍNUO
    """
    # Esta função foi mantida para compatibilidade, mas o novo fluxo
    # utiliza as funções _processar_fluxo_continuo e auxiliares
    pass