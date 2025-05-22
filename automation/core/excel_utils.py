"""
Funções de utilidade para manipulação de planilhas Excel.
"""
import csv
import pandas as pd
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from automation.core.constants import DEFAULT_CONFIG_PATH


def encontrar_coluna_por_data(ws: Worksheet, data):
    """
    Encontra a coluna na planilha que corresponde à data fornecida.
    Verifica tanto formato string quanto objeto datetime.
    
    Args:
        ws (Worksheet): Planilha do openpyxl.
        data (datetime): Data a ser localizada na planilha.
        
    Returns:
        int or None: Número da coluna se encontrada, None caso contrário.
    """
    data_str = data.strftime("%d/%m/%Y")
    
    for col in range(8, ws.max_column + 1):  # Começando da coluna H (8)
        cell_value = ws.cell(row=2, column=col).value
        
        # Verifica se o valor da célula é uma data
        if isinstance(cell_value, datetime):
            # Compara o valor da data
            if cell_value.date() == data.date():
                return col
        elif isinstance(cell_value, str):
            # Tenta converter para datetime se for string
            try:
                # Tenta primeiro o formato DD/MM/YYYY
                cell_date = datetime.strptime(cell_value, "%d/%m/%Y")
                if cell_date.date() == data.date():
                    return col
            except ValueError:
                try:
                    # Tenta o formato YYYY-MM-DD
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    if cell_date.date() == data.date():
                        return col
                except ValueError:
                    # Ignora se não conseguir converter
                    pass
    
    return None

def calcular_producao_planejada(ws: Worksheet, setor_nome: str, coluna: int, linha_limite: int):
    """
    Calcula a produção já planejada para um setor em uma data específica.
    
    Args:
        ws (Worksheet): Planilha do openpyxl.
        setor_nome (str): Nome do setor para calcular a produção.
        coluna (int): Coluna correspondente à data.
        linha_limite (int): Linha limite até onde calcular.
        
    Returns:
        int: Soma da produção planejada para o setor na data.
    """
    valor_planejado = 0
    for row in range(13, linha_limite):
        setor_cell_value = ws.cell(row=row, column=7).value  # Coluna G é a 7ª coluna
        if setor_cell_value == setor_nome:
            cell_value = ws.cell(row=row, column=coluna).value
            valor_planejado += int(cell_value or 0)
            
    return valor_planejado

def obter_limite_producao(ws: Worksheet, linha_limite: int):
    """
    Obtém o limite de produção diário para um setor.
    
    Args:
        ws (Worksheet): Planilha do openpyxl.
        linha_limite (int): Linha onde está o limite do setor.
        
    Returns:
        int: Valor limite de produção diária para o setor.
    """
    try:
        valor_limite_cell = ws.cell(row=linha_limite, column=5).value
        valor_limite_max = int(valor_limite_cell) if valor_limite_cell is not None else 0
        return valor_limite_max
    except (ValueError, TypeError):
        return 0

def obter_carga_producao(config_path=DEFAULT_CONFIG_PATH):
    
    try:
        df = pd.read_csv(config_path, encoding='utf-16')

        # Verifica se as colunas esperadas estão presentes
        if 'PARAMETRO' not in df.columns or 'VALOR' not in df.columns:
            print("⚠ Erro: Colunas 'PARAMETRO' ou 'VALOR' estão ausentes no arquivo CSV.")
            raise ValueError("Colunas obrigatórias ausentes no arquivo CSV.")
        
        config_data = pd.Series(df['VALOR'].values, index=df['PARAMETRO']).to_dict()

        if "CARGA" in config_data:
            return int(config_data['CARGA'])
        else:
            return 100
        
    except FileNotFoundError:
        print(f"Arquivo de configuração não encontrado: {config_path}. Usando valores padrão 100%")
    except ValueError as e:
        print(f"Erro ao converter valores do arquivo de configuração: {e}. Usando valores padrão 100%")
    except Exception as e:
        print(f"Erro inesperado ao carregar o arquivo de configuração: {e}. Usando valores padrão 100%")
    return 100


def atualizar_limites_maximos(config_path=DEFAULT_CONFIG_PATH):
    """
    Obtém o valor de limites máximos do arquivo de configuração csv.

    Args:
        config_path (str): Caminho para o arquivo de configuração CSV.

    Returns:
        list: Lista de valores de limites máximos.
    """
    max_list = []
    standard_list = [5000, 2000, 750, 500, 2000, 350, 800, 500, 1000, 1000]
    max_order_list = [
        'MAX_PCP', 'MAX_SEPARACAO_MP', 'MAX_CORTE_MANUAL', 'MAX_IMPRESSAO',
        'MAX_ESTAMPA', 'MAX_CORTE_LASER','MAX_DISTRIBUICAO', 'MAX_COSTURA', 'MAX_ARREMATE', 'MAX_EMBALAGEM',
    ]
    try:
        df = pd.read_csv(config_path, encoding='utf-16')

        # Verifica se as colunas esperadas estão presentes
        if 'PARAMETRO' not in df.columns or 'VALOR' not in df.columns:
            print("⚠ Erro: Colunas 'PARAMETRO' ou 'VALOR' estão ausentes no arquivo CSV.")
            raise ValueError("Colunas obrigatórias ausentes no arquivo CSV.")

        # Cria um dicionário com os valores do arquivo
        config_data = pd.Series(df['VALOR'].values, index=df['PARAMETRO']).to_dict()

        # Itera na ordem de max_order_list e busca os valores correspondentes
        max_list = []
        for parametro in max_order_list:
            if parametro in config_data:
                try:
                    valor = int(config_data[parametro])
                    max_list.append(valor)
                except ValueError:
                    print(f"⚠ Erro ao converter o valor de '{parametro}' para inteiro. Valor encontrado: '{config_data[parametro]}'")
                    max_list.append(standard_list[max_order_list.index(parametro)])
            else:
                print(f"⚠ Parâmetro '{parametro}' não encontrado no arquivo de configuração. Usando valor padrão {standard_list[max_order_list.index(parametro)]}.")
                max_list.append(standard_list[max_order_list.index(parametro)])

        return max_list

    except FileNotFoundError:
        print(f"Arquivo de configuração não encontrado: {config_path}. Usando valores padrão ({standard_list}).")
    except ValueError as e:
        print(f"Erro ao converter valores do arquivo de configuração: {e}. Usando valores padrão ({standard_list}).")
    except Exception as e:
        print(f"Erro inesperado ao carregar o arquivo de configuração: {e}. Usando valores padrão ({standard_list}).")
    return standard_list

def atualizar_celulas_limite(ws: Worksheet, max_list: list):
    """
    Atualiza os valores das células na faixa [E3:E11] com os valores da max_list.

    Args:
        ws (Worksheet): Planilha do openpyxl.
        max_list (list): Lista de valores para atualizar as células.
    """
    # Verifica se a lista tem exatamente 9 valores (correspondente a E3:E11)
    if len(max_list) != 10:
        raise ValueError("A lista max_list deve conter exatamente 10 valores para atualizar as células [E3:E11].")

    # Itera sobre os índices e valores da max_list
    for i, valor in enumerate(max_list):
        # A linha começa em 3 (E3) e vai até 12 (E12)
        linha = 3 + i
        coluna = 5  # Coluna E é a 5ª coluna
        ws.cell(row=linha, column=coluna, value=valor)
