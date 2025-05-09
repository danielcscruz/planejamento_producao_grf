from openpyxl import load_workbook
import pandas as pd
from automation.fill_production import preencher_producao  

def criar_novo_plano(df_priorizado: pd.DataFrame):
    total = len(df_priorizado)
    arquivo_path = "model/planejamento.xlsx"

    wb = load_workbook(arquivo_path)
    ws = wb.active

    for index, row in df_priorizado.iterrows():
        linha = 12
        while ws.cell(row=linha, column=1).value:
            linha += 1

        # Pegar valores da tabela
        pedido = row.get("PEDIDO")
        entrega = row.get("ENTREGA")
        cliente = row.get("CLIENTE")
        produto = row.get("PRODUTO")
        quantidade= row.get("QUANTIDADE")
        corte = row.get("TIPO DE CORTE")
        
        # Preenche os dados na planilha
        ws.cell(row=linha, column=1, value=pedido)
        ws.cell(row=linha, column=2, value=entrega)
        ws.cell(row=linha, column=3, value=cliente)
        ws.cell(row=linha, column=4, value=produto)
        ws.cell(row=linha, column=5, value=quantidade)


        if not corte:
            print(f"[Linha {index}] TIPO DE CORTE não especificado. Pulando.")
            continue

        try:
            salvar = (index == total - 1)  # só salva no último
            print(f"[Linha {index}] Iniciando preenchimento para tipo de corte: {corte}")

            preencher_producao(ws, quantidade=quantidade, setor="PCP", linha=linha, calendario_path="data/_CALENDARIO.csv", planilha_path=arquivo_path, workbook=wb, corte=corte, salvar=salvar)

            print(f"[Linha {index}] Preenchimento concluído.\n")
        except Exception as e:
            print(f"[Linha {index}] Erro ao preencher produção: {e}")
